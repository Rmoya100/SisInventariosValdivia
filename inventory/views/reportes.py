from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from ..decorators import permission_or_admin
# Mantener compatibilidad: definir `permission_required` que acepta los mismos kwargs pero delega a permission_or_admin
def permission_required(perm, login_url=None, raise_exception=False):
    return permission_or_admin(perm)
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q, Sum, Max
from django.db.models.functions import Coalesce
from openpyxl import Workbook
from ..models import (
    OrdenCompra, Ingreso, Salida, Producto, Proveedor, Categoria, Empresa, Transferencia,
    StockProyecto, DetalleIngreso, DetalleSalida, Herramienta, Maquinaria, Proyecto,
    Bodega, DetalleTransferencia, ModuloTorre, Gasto, Fase
)


@login_required
@permission_required('inventory.view_ordencompra', raise_exception=True)
def reporte_compras_pdf(request):
    q = request.GET.get('q')
    proyecto_id = request.GET.get('proyecto')
    estado = request.GET.get('estado')
    
    ordenes = OrdenCompra.objects.all().prefetch_related('detalles__producto')
    
    if request.user.proyecto and not request.user.es_admin:
        ordenes = ordenes.filter(proyecto=request.user.proyecto)
    elif proyecto_id:
        ordenes = ordenes.filter(proyecto_id=proyecto_id)
        
    if q:
        ordenes = ordenes.filter(Q(proveedor__nombre__icontains=q) | Q(numCompra__icontains=q))
        
    if estado:
        ordenes = ordenes.filter(estado=estado)
        
    titulo_reporte = 'Reporte de Órdenes de Compra'
    filtros = []
    if proyecto_id:
        proj = Proyecto.objects.filter(pk=proyecto_id).first()
        if proj:
            filtros.append(f'Proyecto: {proj.nombre}')
    if estado:
        filtros.append(f'Estado: {estado}')
    if q:
        filtros.append(f'Búsqueda: {q}')
        
    if filtros:
        titulo_reporte += f" ({', '.join(filtros)})"
        
    empresa = Empresa.objects.first()
    context = {
        'ordenes': ordenes,
        'titulo_reporte': titulo_reporte,
        'empresa': empresa,
        'pendientes': ordenes.filter(estado='PENDIENTE').count(),
        'parciales': ordenes.filter(estado='RECEPCION PARCIAL').count(),
        'completadas': ordenes.filter(estado='RECEPCION OK').count(),
    }
    return render(request, 'inventory/reportes/compras.html', context)

@login_required
@permission_required('inventory.view_ordencompra', raise_exception=True)
def reporte_oc_pendientes_pdf(request):
    ordenes = OrdenCompra.objects.filter(estado__in=['PENDIENTE', 'RECEPCION PARCIAL']).prefetch_related('detalles__producto')
    empresa = Empresa.objects.first()
    context = {
        'ordenes': ordenes,
        'titulo_reporte': 'Reporte de Órdenes Pendientes / Parciales',
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/compras.html', context)

@login_required
@permission_required('inventory.view_ingreso', raise_exception=True)
def reporte_ingresos_pdf(request):
    empresa = Empresa.objects.first()
    ingresos = Ingreso.objects.all().prefetch_related('detalles__producto')
    total_valorizado = DetalleIngreso.objects.aggregate(total=Sum('subtotal'))['total'] or 0
    context = {
        'ingresos': ingresos,
        'empresa': empresa,
        'total_valorizado': total_valorizado,
    }
    return render(request, 'inventory/reportes/ingresos.html', context)


@login_required
@permission_required('inventory.view_ingreso', raise_exception=True)
def reporte_ingreso_detail_pdf(request, pk):
    ingreso = get_object_or_404(Ingreso, pk=pk)
    empresa = Empresa.objects.first()
    total_ingreso = ingreso.detalles.aggregate(total=Sum('subtotal'))['total'] or 0
    context = {
        'ingreso': ingreso,
        'empresa': empresa,
        'total_ingreso': total_ingreso,
    }
    return render(request, 'inventory/reportes/ingreso_detail.html', context)

@login_required
@permission_required('inventory.view_salida', raise_exception=True)
def reporte_salidas_pdf(request):
    empresa = Empresa.objects.first()
    context = {
        'salidas': Salida.objects.all().prefetch_related('detalles__producto'),
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/salidas.html', context)

@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def reporte_productos_pdf(request):
    empresa = Empresa.objects.first()
    context = {
        'productos': Producto.objects.all(),
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/productos.html', context)

@login_required
@permission_required('inventory.view_proveedor', raise_exception=True)
def reporte_proveedores_pdf(request):
    empresa = Empresa.objects.first()
    context = {
        'proveedores': Proveedor.objects.all(),
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/proveedores.html', context)

@login_required
@permission_required('inventory.view_categoria', raise_exception=True)
def reporte_categorias_pdf(request):
    empresa = Empresa.objects.first()
    context = {
        'categorias': Categoria.objects.all(),
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/categorias.html', context)

@login_required
@permission_required('inventory.view_transferencia', raise_exception=True)
def reporte_transferencia_pdf(request, pk):
    transferencia = get_object_or_404(Transferencia, pk=pk)
    empresa = Empresa.objects.first()
    context = {
        't': transferencia,
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/transferencia.html', context)

@login_required
@permission_required('inventory.view_ordencompra', raise_exception=True)
def reporte_compras_list_view(request):
    q = request.GET.get('q')
    proyecto_id = request.GET.get('proyecto')
    estado = request.GET.get('estado')
    
    ordenes = OrdenCompra.objects.all().select_related('proveedor', 'proyecto').prefetch_related('detalles__producto')
    if request.user.proyecto and not request.user.es_admin:
        ordenes = ordenes.filter(proyecto=request.user.proyecto)
    elif proyecto_id:
        ordenes = ordenes.filter(proyecto_id=proyecto_id)
        
    if q:
        ordenes = ordenes.filter(Q(proveedor__nombre__icontains=q) | Q(numCompra__icontains=q))
        
    if estado:
        ordenes = ordenes.filter(estado=estado)
        
    paginator = Paginator(ordenes, 25)
    page = request.GET.get('page', 1)
    ordenes_page = paginator.get_page(page)

    context = {
        'ordenes': ordenes_page,
        'proyectos': Proyecto.objects.all(),
        'q': q or '',
        'proyecto_id': int(proyecto_id) if proyecto_id else '',
        'estado': estado or '',
        'is_admin': request.user.es_admin,
        'paginator': paginator,
        'page_obj': ordenes_page,
    }
    return render(request, 'inventory/reporte_compras_list.html', context)

@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def reporte_stock_bodega_list_view(request):
    q = request.GET.get('q', '').strip()
    proyecto_id = request.GET.get('proyecto')
    stocks = StockProyecto.objects.select_related('producto__categoria', 'producto__unidad_medida', 'proyecto').all()
    if q:
        stocks = stocks.filter(producto__nombre__icontains=q)
    if proyecto_id:
        stocks = stocks.filter(proyecto_id=proyecto_id)

    paginator = Paginator(stocks, 50)
    page = request.GET.get('page', 1)
    stocks_page = paginator.get_page(page)

    context = {
        'stocks': stocks_page,
        'proyectos': Proyecto.objects.all(),
        'q': q,
        'proyecto_id': int(proyecto_id) if proyecto_id else '',
        'paginator': paginator,
        'page_obj': stocks_page,
    }
    return render(request, 'inventory/reporte_stock_bodega_list.html', context)

@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def reporte_stock_bodega_pdf(request):
    q = request.GET.get('q', '').strip()
    proyecto_id = request.GET.get('proyecto')
    stocks = StockProyecto.objects.select_related('producto__categoria', 'producto__unidad_medida', 'proyecto').all()
    if q:
        stocks = stocks.filter(producto__nombre__icontains=q)
    if proyecto_id:
        stocks = stocks.filter(proyecto_id=proyecto_id)

    empresa = Empresa.objects.first()
    context = {
        'stocks': stocks,
        'titulo_reporte': 'Reporte de Stock por Bodega/Proyecto',
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/stock_bodega.html', context)

@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def exportar_stock_bodega_excel(request):
    q = request.GET.get('q', '').strip()
    proyecto_id = request.GET.get('proyecto')
    stocks = StockProyecto.objects.select_related('producto__categoria', 'producto__unidad_medida', 'proyecto').all()
    if q:
        stocks = stocks.filter(producto__nombre__icontains=q)
    if proyecto_id:
        stocks = stocks.filter(proyecto_id=proyecto_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock por Bodega'
    ws.append(['Producto', 'Categoría', 'Unidad de Medida', 'Proyecto/Bodega', 'Cantidad'])
    for stock in stocks.iterator():
        unidad = stock.producto.unidad_medida.abreviatura if stock.producto.unidad_medida else ''
        ws.append([stock.producto.nombre, stock.producto.categoria.nombre, unidad, stock.proyecto.nombre if stock.proyecto else 'Sin Proyecto', stock.cantidad])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stock_por_bodega.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def reporte_movimiento_general_list_view(request):
    q = request.GET.get('q', '').strip()
    productos = Producto.objects.select_related('categoria', 'unidad_medida').all()
    if q:
        productos = productos.filter(nombre__icontains=q)

    ingresos = DetalleIngreso.objects.values('producto').annotate(total_ing=Coalesce(Sum('cantidad'), 0))
    salidas = DetalleSalida.objects.values('producto').annotate(total_sal=Coalesce(Sum('cantidad'), 0))
    ingreso_map = {item['producto']: item['total_ing'] for item in ingresos}
    salida_map = {item['producto']: item['total_sal'] for item in salidas}

    movimientos = []
    for producto in productos:
        total_ing = ingreso_map.get(producto.pk, 0)
        total_sal = salida_map.get(producto.pk, 0)
        movimientos.append({
            'producto': producto,
            'categoria': producto.categoria.nombre,
            'unidad_medida': producto.unidad_medida,
            'stock_inicial': producto.stock_inicial,
            'ingresos': total_ing,
            'salidas': total_sal,
            'stock_actual': producto.stock_inicial + total_ing - total_sal,
        })

    paginator = Paginator(movimientos, 50)
    page = request.GET.get('page', 1)
    movimientos_page = paginator.get_page(page)

    context = {
        'movimientos': movimientos_page,
        'q': q,
        'paginator': paginator,
        'page_obj': movimientos_page,
    }
    return render(request, 'inventory/reporte_movimiento_general_list.html', context)

@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def reporte_movimiento_general_pdf(request):
    q = request.GET.get('q', '').strip()
    productos = Producto.objects.select_related('categoria', 'unidad_medida').all()
    if q:
        productos = productos.filter(nombre__icontains=q)

    ingresos = DetalleIngreso.objects.values('producto').annotate(total_ing=Coalesce(Sum('cantidad'), 0))
    salidas = DetalleSalida.objects.values('producto').annotate(total_sal=Coalesce(Sum('cantidad'), 0))
    ingreso_map = {item['producto']: item['total_ing'] for item in ingresos}
    salida_map = {item['producto']: item['total_sal'] for item in salidas}

    movimientos = []
    for producto in productos:
        total_ing = ingreso_map.get(producto.pk, 0)
        total_sal = salida_map.get(producto.pk, 0)
        movimientos.append({
            'producto': producto,
            'categoria': producto.categoria.nombre,
            'unidad_medida': producto.unidad_medida,
            'stock_inicial': producto.stock_inicial,
            'ingresos': total_ing,
            'salidas': total_sal,
            'stock_actual': producto.stock_inicial + total_ing - total_sal,
        })

    empresa = Empresa.objects.first()
    context = {
        'movimientos': movimientos,
        'titulo_reporte': 'Reporte General de Movimiento de Inventario',
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/movimiento_general.html', context)

@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def exportar_movimiento_general_excel(request):
    q = request.GET.get('q', '').strip()
    productos = Producto.objects.select_related('categoria', 'unidad_medida').all()
    if q:
        productos = productos.filter(nombre__icontains=q)

    ingresos = DetalleIngreso.objects.values('producto').annotate(total_ing=Coalesce(Sum('cantidad'), 0))
    salidas = DetalleSalida.objects.values('producto').annotate(total_sal=Coalesce(Sum('cantidad'), 0))
    ingreso_map = {item['producto']: item['total_ing'] for item in ingresos}
    salida_map = {item['producto']: item['total_sal'] for item in salidas}

    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimiento General'
    ws.append(['Producto', 'Categoría', 'Unidad de Medida', 'Stock Inicial', 'Ingresado', 'Salida', 'Stock Actual'])
    for producto in productos.iterator():
        total_ing = ingreso_map.get(producto.pk, 0)
        total_sal = salida_map.get(producto.pk, 0)
        unidad = producto.unidad_medida.abreviatura if producto.unidad_medida else ''
        ws.append([
            producto.nombre,
            producto.categoria.nombre,
            unidad,
            producto.stock_inicial,
            total_ing,
            total_sal,
            producto.stock_inicial + total_ing - total_sal,
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=movimiento_general.xlsx'
    wb.save(response)
    return response


@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def exportar_productos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(['Código', 'Nombre', 'Categoría', 'Stock Inicial', 'Stock Real'])
    for obj in Producto.objects.all().select_related('categoria').iterator():
        ws.append([obj.cod_prod, obj.nombre, obj.categoria.nombre, obj.stock_inicial, obj.stock_actual])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('inventory.view_proveedor', raise_exception=True)
def exportar_proveedores_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    ws.append(['Nombre', 'Dirección', 'Teléfono', 'Contacto', 'Correo'])
    for obj in Proveedor.objects.all().iterator():
        ws.append([obj.nombre, obj.direccion, obj.telefono, obj.contacto, obj.correo])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('inventory.view_categoria', raise_exception=True)
def exportar_categorias_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    ws.append(['ID', 'Nombre'])
    for obj in Categoria.objects.all().iterator():
        ws.append([obj.idCategoria, obj.nombre])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=categorias.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('inventory.view_ordencompra', raise_exception=True)
def exportar_compras_excel(request):
    q = request.GET.get('q')
    proyecto_id = request.GET.get('proyecto')
    estado = request.GET.get('estado')
    
    ordenes = OrdenCompra.objects.all().select_related('proveedor')
    
    if request.user.proyecto and not request.user.es_admin:
        ordenes = ordenes.filter(proyecto=request.user.proyecto)
    elif proyecto_id:
        ordenes = ordenes.filter(proyecto_id=proyecto_id)
        
    if q:
        ordenes = ordenes.filter(Q(proveedor__nombre__icontains=q) | Q(numCompra__icontains=q))
        
    if estado:
        ordenes = ordenes.filter(estado=estado)
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes de Compra"
    ws.append(['N° Compra', 'Proveedor', 'Fecha', 'Forma de Pago', 'Estado'])
    for obj in ordenes.iterator():
        ws.append([obj.numCompra, obj.proveedor.nombre, obj.fecha_compra.strftime('%d/%m/%Y'), obj.forma_de_pago, obj.estado])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ordenes_compra.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('inventory.view_ingreso', raise_exception=True)
def exportar_ingresos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingresos"
    ws.append(['N° Ingreso', 'Fecha', 'N° Compra', 'Tipo Doc', 'Num Doc'])
    for obj in Ingreso.objects.all().select_related('orden_compra').iterator():
        ws.append([obj.numIngreso, obj.fecha.strftime('%d/%m/%Y'), obj.orden_compra.numCompra, obj.tipo_documento, obj.num_documento])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ingresos.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('inventory.view_salida', raise_exception=True)
def exportar_salidas_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Salidas"
    ws.append(['N° Salida', 'Fecha', 'Módulo/Torre', 'Solicitante'])
    for obj in Salida.objects.select_related('modulo_torre', 'solicitante').all().iterator():
        ws.append([
            obj.numSalida,
            obj.fecha.strftime('%d/%m/%Y'),
            str(obj.modulo_torre) if obj.modulo_torre else '',
            str(obj.solicitante) if obj.solicitante else '',
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=salidas.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('inventory.view_herramienta', raise_exception=True)
def reporte_herramientas_pdf(request):
    q = request.GET.get('q')
    herramientas = Herramienta.objects.select_related('bodega_actual').all()
    if q:
        herramientas = herramientas.filter(Q(nomHerramienta__icontains=q) | Q(codigo__icontains=q) | Q(marca__icontains=q))
    empresa = Empresa.objects.first()
    context = {
        'herramientas': herramientas,
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/herramientas.html', context)

@login_required
@permission_required('inventory.view_herramienta', raise_exception=True)
def exportar_herramientas_excel(request):
    q = request.GET.get('q')
    herramientas = Herramienta.objects.select_related('bodega_actual').all()
    if q:
        herramientas = herramientas.filter(Q(nomHerramienta__icontains=q) | Q(codigo__icontains=q) | Q(marca__icontains=q))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Herramientas"
    ws.append(['Código', 'Nombre', 'Marca', 'Estado', 'Bodega Actual', 'En Reparación'])
    for obj in herramientas.iterator():
        bodega_str = obj.bodega_actual.nombre if obj.bodega_actual else "Sin Asignar"
        ws.append([obj.codigo, obj.nomHerramienta, obj.marca, obj.get_estado_display(), bodega_str, "Sí" if obj.en_reparacion else "No"])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=herramientas.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('inventory.view_maquinaria', raise_exception=True)
def reporte_maquinarias_pdf(request):
    q = request.GET.get('q')
    maquinarias = Maquinaria.objects.select_related('bodega_actual').all()
    if q:
        maquinarias = maquinarias.filter(Q(marca__icontains=q) | Q(patente_o_codigo__icontains=q) | Q(modelo__icontains=q))
    empresa = Empresa.objects.first()
    context = {
        'maquinarias': maquinarias,
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/maquinarias.html', context)

@login_required
@permission_required('inventory.view_maquinaria', raise_exception=True)
def exportar_maquinarias_excel(request):
    q = request.GET.get('q')
    maquinarias = Maquinaria.objects.select_related('bodega_actual').all()
    if q:
        maquinarias = maquinarias.filter(Q(marca__icontains=q) | Q(patente_o_codigo__icontains=q) | Q(modelo__icontains=q))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Maquinarias"
    ws.append(['Patente/Código', 'Tipo Máquina', 'Marca', 'Modelo', 'Tipo Control', 'Valor Actual', 'Alerta Mant.', 'Bodega Actual'])
    for obj in maquinarias.iterator():
        bodega_str = obj.bodega_actual.nombre if obj.bodega_actual else "Sin Asignar"
        alerta = "Sí" if obj.alerta_mantenimiento() else "No"
        ws.append([
            obj.patente_o_codigo, obj.get_tipo_maquina_display(), obj.marca, 
            obj.modelo or "", obj.get_tipo_control_display(), obj.valor_actual, 
            alerta, bodega_str
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=maquinarias.xlsx'
    wb.save(response)
    return response


@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def reporte_movimientos_producto_list_view(request):
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    # Determinar si el usuario es administrador o staff (ambos pueden seleccionar bodega)
    is_admin = request.user.es_admin or request.user.is_staff
    
    # Obtener bodegas activas para el selector de admin/staff
    bodegas = Bodega.objects.filter(activo=True).select_related('proyecto') if is_admin else None
    
    # Bodega del usuario (estándar)
    bodega_usuario = None
    if not is_admin:
        bodega_usuario = request.user.get_bodega_movimiento()

    # Parámetros del request
    producto_id = request.GET.get('producto')
    bodega_id = request.GET.get('bodega')

    # Resolver bodega y producto seleccionados
    producto_seleccionado = None
    bodega_seleccionada = None

    if producto_id:
        producto_seleccionado = Producto.objects.filter(pk=producto_id).first()

    if is_admin:
        if bodega_id:
            bodega_seleccionada = Bodega.objects.filter(pk=bodega_id).first()
    else:
        bodega_seleccionada = bodega_usuario

    movimientos = []
    stock_real = 0
    
    # Si ambos están seleccionados, buscar movimientos
    if producto_seleccionado and bodega_seleccionada:
        # 1. Ingresos
        ingresos = DetalleIngreso.objects.filter(
            producto=producto_seleccionado,
            ingreso__bodega=bodega_seleccionada
        ).select_related('ingreso__orden_compra')
        
        for det in ingresos:
            movimientos.append({
                'fecha': det.ingreso.fecha,
                'tipo': 'Ingreso',
                'referencia': f'Ingreso N° {det.ingreso.numIngreso}',
                'detalle_ref': f'OC {det.ingreso.orden_compra.numCompra}' if det.ingreso.orden_compra else 'Ingreso Directo',
                'cantidad': det.cantidad,
            })
            
        # 2. Salidas
        salidas = DetalleSalida.objects.filter(
            producto=producto_seleccionado,
            salida__bodega=bodega_seleccionada
        ).select_related('salida__modulo_torre', 'salida__solicitante')
        
        for det in salidas:
            ref_parts = []
            if det.salida.modulo_torre:
                ref_parts.append(f"Módulo: {det.salida.modulo_torre.nombre}")
            if det.salida.solicitante:
                ref_parts.append(f"Solicitante: {det.salida.solicitante}")
            movimientos.append({
                'fecha': det.salida.fecha,
                'tipo': 'Salida',
                'referencia': f'Salida N° {det.salida.numSalida}',
                'detalle_ref': ' - '.join(ref_parts) if ref_parts else 'Salida',
                'cantidad': -det.cantidad,
            })
            
        # 3. GDI Despachos (Transferencias as origin)
        if bodega_seleccionada.proyecto:
            despachos = DetalleTransferencia.objects.filter(
                producto=producto_seleccionado,
                transferencia__proyecto_origen=bodega_seleccionada.proyecto
            ).select_related('transferencia__proyecto_destino')
            
            for det in despachos:
                movimientos.append({
                    'fecha': det.transferencia.fecha_despacho,
                    'tipo': 'GDI (Despacho)',
                    'referencia': f'GDI N° {det.transferencia.idTransferencia}',
                    'detalle_ref': f'Destino: {det.transferencia.proyecto_destino.nombre}',
                    'cantidad': -det.cantidad_enviada,
                })
                
            # 4. GDI Recepciones (Transferencias as destination)
            recepciones = DetalleTransferencia.objects.filter(
                producto=producto_seleccionado,
                transferencia__proyecto_destino=bodega_seleccionada.proyecto,
                cantidad_recibida__gt=0
            ).select_related('transferencia__proyecto_origen')
            
            for det in recepciones:
                movimientos.append({
                    'fecha': det.transferencia.fecha_despacho,
                    'tipo': 'GDI (Recepción)',
                    'referencia': f'GDI N° {det.transferencia.idTransferencia}',
                    'detalle_ref': f'Origen: {det.transferencia.proyecto_origen.nombre}',
                    'cantidad': det.cantidad_recibida,
                })
                
        # Ordenar cronológicamente para calcular stock acumulado
        movimientos.sort(key=lambda m: (m['fecha'], m['tipo']))
        
        stock_acumulado = 0
        for mov in movimientos:
            stock_acumulado += mov['cantidad']
            mov['stock_acumulado'] = stock_acumulado
            
        # Revertir para mostrar primero el más reciente
        movimientos.reverse()
        
        # Stock real actual
        if bodega_seleccionada.proyecto:
            sp = StockProyecto.objects.filter(producto=producto_seleccionado, proyecto=bodega_seleccionada.proyecto).first()
            if sp:
                stock_real = sp.cantidad

    # Paginación
    paginator = Paginator(movimientos, 50)
    page = request.GET.get('page', 1)
    movimientos_page = paginator.get_page(page)

    context = {
        'productos': productos,
        'bodegas': bodegas,
        'is_admin': is_admin,
        'bodega_usuario': bodega_usuario,
        'producto_seleccionado': producto_seleccionado,
        'bodega_seleccionada': bodega_seleccionada,
        'producto_id': int(producto_id) if producto_id else '',
        'bodega_id': int(bodega_id) if bodega_id else '',
        'movimientos': movimientos_page,
        'stock_real': stock_real,
        'paginator': paginator,
        'page_obj': movimientos_page,
        'total_movimientos': len(movimientos),
    }
    
    return render(request, 'inventory/reporte_movimientos_producto_list.html', context)


@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def reporte_movimientos_producto_pdf(request):
    is_admin = request.user.es_admin or request.user.is_staff
    producto_id = request.GET.get('producto')
    bodega_id = request.GET.get('bodega')

    producto_seleccionado = None
    bodega_seleccionada = None

    if producto_id:
        producto_seleccionado = Producto.objects.filter(pk=producto_id).first()

    if is_admin:
        if bodega_id:
            bodega_seleccionada = Bodega.objects.filter(pk=bodega_id).first()
    else:
        bodega_seleccionada = request.user.get_bodega_movimiento()

    movimientos = []
    stock_real = 0

    if producto_seleccionado and bodega_seleccionada:
        # 1. Ingresos
        ingresos = DetalleIngreso.objects.filter(
            producto=producto_seleccionado,
            ingreso__bodega=bodega_seleccionada
        ).select_related('ingreso__orden_compra')
        for det in ingresos:
            movimientos.append({
                'fecha': det.ingreso.fecha,
                'tipo': 'Ingreso',
                'referencia': f'Ingreso N° {det.ingreso.numIngreso}',
                'detalle_ref': f'OC {det.ingreso.orden_compra.numCompra}' if det.ingreso.orden_compra else 'Ingreso Directo',
                'cantidad': det.cantidad,
            })
            
        # 2. Salidas
        salidas = DetalleSalida.objects.filter(
            producto=producto_seleccionado,
            salida__bodega=bodega_seleccionada
        ).select_related('salida__modulo_torre', 'salida__solicitante')
        for det in salidas:
            ref_parts = []
            if det.salida.modulo_torre:
                ref_parts.append(f"Módulo: {det.salida.modulo_torre.nombre}")
            if det.salida.solicitante:
                ref_parts.append(f"Solicitante: {det.salida.solicitante}")
            movimientos.append({
                'fecha': det.salida.fecha,
                'tipo': 'Salida',
                'referencia': f'Salida N° {det.salida.numSalida}',
                'detalle_ref': ' - '.join(ref_parts) if ref_parts else 'Salida',
                'cantidad': -det.cantidad,
            })
            
        # 3. GDI Despachos
        if bodega_seleccionada.proyecto:
            despachos = DetalleTransferencia.objects.filter(
                producto=producto_seleccionado,
                transferencia__proyecto_origen=bodega_seleccionada.proyecto
            ).select_related('transferencia__proyecto_destino')
            for det in despachos:
                movimientos.append({
                    'fecha': det.transferencia.fecha_despacho,
                    'tipo': 'GDI (Despacho)',
                    'referencia': f'GDI N° {det.transferencia.idTransferencia}',
                    'detalle_ref': f'Destino: {det.transferencia.proyecto_destino.nombre}',
                    'cantidad': -det.cantidad_enviada,
                })
                
            # 4. GDI Recepciones
            recepciones = DetalleTransferencia.objects.filter(
                producto=producto_seleccionado,
                transferencia__proyecto_destino=bodega_seleccionada.proyecto,
                cantidad_recibida__gt=0
            ).select_related('transferencia__proyecto_origen')
            for det in recepciones:
                movimientos.append({
                    'fecha': det.transferencia.fecha_despacho,
                    'tipo': 'GDI (Recepción)',
                    'referencia': f'GDI N° {det.transferencia.idTransferencia}',
                    'detalle_ref': f'Origen: {det.transferencia.proyecto_origen.nombre}',
                    'cantidad': det.cantidad_recibida,
                })
                
        movimientos.sort(key=lambda m: (m['fecha'], m['tipo']))
        stock_acumulado = 0
        for mov in movimientos:
            stock_acumulado += mov['cantidad']
            mov['stock_acumulado'] = stock_acumulado
            
        movimientos.reverse()
        
        if bodega_seleccionada.proyecto:
            sp = StockProyecto.objects.filter(producto=producto_seleccionado, proyecto=bodega_seleccionada.proyecto).first()
            if sp:
                stock_real = sp.cantidad

    empresa = Empresa.objects.first()
    titulo_reporte = f"Reporte de Movimientos de {producto_seleccionado.nombre if producto_seleccionado else ''}"
    if bodega_seleccionada:
        titulo_reporte += f" - Bodega: {bodega_seleccionada.nombre}"

    context = {
        'movimientos': movimientos,
        'producto_seleccionado': producto_seleccionado,
        'bodega_seleccionada': bodega_seleccionada,
        'stock_real': stock_real,
        'titulo_reporte': titulo_reporte,
        'empresa': empresa,
    }
    return render(request, 'inventory/reportes/movimientos_producto.html', context)


@login_required
@permission_required('inventory.view_producto', raise_exception=True)
def exportar_movimientos_producto_excel(request):
    is_admin = request.user.es_admin or request.user.is_staff
    producto_id = request.GET.get('producto')
    bodega_id = request.GET.get('bodega')

    producto_seleccionado = None
    bodega_seleccionada = None

    if producto_id:
        producto_seleccionado = Producto.objects.filter(pk=producto_id).first()

    if is_admin:
        if bodega_id:
            bodega_seleccionada = Bodega.objects.filter(pk=bodega_id).first()
    else:
        bodega_seleccionada = request.user.get_bodega_movimiento()

    movimientos = []
    stock_real = 0

    if producto_seleccionado and bodega_seleccionada:
        # 1. Ingresos
        ingresos = DetalleIngreso.objects.filter(
            producto=producto_seleccionado,
            ingreso__bodega=bodega_seleccionada
        ).select_related('ingreso__orden_compra')
        for det in ingresos:
            movimientos.append({
                'fecha': det.ingreso.fecha,
                'tipo': 'Ingreso',
                'referencia': f'Ingreso N° {det.ingreso.numIngreso}',
                'detalle_ref': f'OC {det.ingreso.orden_compra.numCompra}' if det.ingreso.orden_compra else 'Ingreso Directo',
                'cantidad': det.cantidad,
            })
            
        # 2. Salidas
        salidas = DetalleSalida.objects.filter(
            producto=producto_seleccionado,
            salida__bodega=bodega_seleccionada
        ).select_related('salida__modulo_torre', 'salida__solicitante')
        for det in salidas:
            ref_parts = []
            if det.salida.modulo_torre:
                ref_parts.append(f"Módulo: {det.salida.modulo_torre.nombre}")
            if det.salida.solicitante:
                ref_parts.append(f"Solicitante: {det.salida.solicitante}")
            movimientos.append({
                'fecha': det.salida.fecha,
                'tipo': 'Salida',
                'referencia': f'Salida N° {det.salida.numSalida}',
                'detalle_ref': ' - '.join(ref_parts) if ref_parts else 'Salida',
                'cantidad': -det.cantidad,
            })
            
        # 3. GDI Despachos
        if bodega_seleccionada.proyecto:
            despachos = DetalleTransferencia.objects.filter(
                producto=producto_seleccionado,
                transferencia__proyecto_origen=bodega_seleccionada.proyecto
            ).select_related('transferencia__proyecto_destino')
            for det in despachos:
                movimientos.append({
                    'fecha': det.transferencia.fecha_despacho,
                    'tipo': 'GDI (Despacho)',
                    'referencia': f'GDI N° {det.transferencia.idTransferencia}',
                    'detalle_ref': f'Destino: {det.transferencia.proyecto_destino.nombre}',
                    'cantidad': -det.cantidad_enviada,
                })
                
            # 4. GDI Recepciones
            recepciones = DetalleTransferencia.objects.filter(
                producto=producto_seleccionado,
                transferencia__proyecto_destino=bodega_seleccionada.proyecto,
                cantidad_recibida__gt=0
            ).select_related('transferencia__proyecto_origen')
            for det in recepciones:
                movimientos.append({
                    'fecha': det.transferencia.fecha_despacho,
                    'tipo': 'GDI (Recepción)',
                    'referencia': f'GDI N° {det.transferencia.idTransferencia}',
                    'detalle_ref': f'Origen: {det.transferencia.proyecto_origen.nombre}',
                    'cantidad': det.cantidad_recibida,
                })
                
        movimientos.sort(key=lambda m: (m['fecha'], m['tipo']))
        stock_acumulado = 0
        for mov in movimientos:
            stock_acumulado += mov['cantidad']
            mov['stock_acumulado'] = stock_acumulado
            
        movimientos.reverse()
        
        if bodega_seleccionada.proyecto:
            sp = StockProyecto.objects.filter(producto=producto_seleccionado, proyecto=bodega_seleccionada.proyecto).first()
            if sp:
                stock_real = sp.cantidad

    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimientos Producto'
    
    # Encabezados informativos
    ws.append(['Reporte de Movimientos de Producto'])
    ws.append([f'Producto: {producto_seleccionado.nombre if producto_seleccionado else "No Seleccionado"}'])
    ws.append([f'Bodega: {bodega_seleccionada.nombre if bodega_seleccionada else "No Seleccionada"}'])
    ws.append([f'Stock Real Actual: {stock_real}'])
    ws.append([]) # Fila vacía
    
    ws.append(['Fecha', 'Tipo', 'Referencia', 'Detalle', 'Cantidad', 'Stock Acumulado'])
    for mov in movimientos:
        ws.append([
            mov['fecha'].strftime('%d/%m/%Y') if hasattr(mov['fecha'], 'strftime') else str(mov['fecha']),
            mov['tipo'],
            mov['referencia'],
            mov['detalle_ref'],
            mov['cantidad'],
            mov['stock_acumulado']
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"movimientos_{producto_seleccionado.nombre.replace(' ', '_') if producto_seleccionado else 'producto'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
@permission_required('inventory.view_salida', raise_exception=True)
def reporte_gasto_modulo_torre_list_view(request):
    modulo_id = request.GET.get('modulo')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    modulos = ModuloTorre.objects.all().order_by('nombre')
    salidas_det = DetalleSalida.objects.select_related('salida__modulo_torre', 'producto')
    
    if request.user.proyecto and not request.user.es_admin:
        salidas_det = salidas_det.filter(salida__proyecto=request.user.proyecto)
        
    if modulo_id:
        if modulo_id == 'sin_modulo':
            salidas_det = salidas_det.filter(salida__modulo_torre__isnull=True)
        else:
            salidas_det = salidas_det.filter(salida__modulo_torre_id=modulo_id)
            
    if fecha_inicio:
        salidas_det = salidas_det.filter(salida__fecha__gte=fecha_inicio)
    if fecha_fin:
        salidas_det = salidas_det.filter(salida__fecha__lte=fecha_fin)
        
    precios_dict = {}
    ultimos_ingresos = DetalleIngreso.objects.values('producto').annotate(max_id=Max('idDetalle'))
    ids_ingresos = [i['max_id'] for i in ultimos_ingresos]
    precios_qs = DetalleIngreso.objects.filter(idDetalle__in=ids_ingresos).values('producto_id', 'precio')
    for p in precios_qs:
        precios_dict[p['producto_id']] = float(p['precio'])
        
    context = {
        'modulos': modulos,
        'modulo_id': modulo_id or '',
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
    }
    
    if modulo_id:
        productos_gasto = {}
        for d in salidas_det:
            prod_id = d.producto_id
            precio_unitario = precios_dict.get(prod_id, 0.0)
            subtotal = float(d.cantidad) * precio_unitario
            
            if prod_id not in productos_gasto:
                productos_gasto[prod_id] = {
                    'producto': d.producto,
                    'cantidad_total': 0,
                    'precio_unitario': precio_unitario,
                    'total_valorizado': 0.0
                }
            productos_gasto[prod_id]['cantidad_total'] += d.cantidad
            productos_gasto[prod_id]['total_valorizado'] += subtotal
            
        productos_gasto_list = sorted(productos_gasto.values(), key=lambda x: x['total_valorizado'], reverse=True)
        total_modulo = sum(p['total_valorizado'] for p in productos_gasto_list)
        total_cantidades = sum(p['cantidad_total'] for p in productos_gasto_list)
        
        paginator = Paginator(productos_gasto_list, 50)
        page = request.GET.get('page', 1)
        page_obj = paginator.get_page(page)
        
        if modulo_id == 'sin_modulo':
            modulo_nombre = 'Sin Módulo / Torre'
        else:
            modulo_obj = get_object_or_404(ModuloTorre, pk=modulo_id)
            modulo_nombre = modulo_obj.nombre
            
        # Top 10 productos para gráfico de barras
        top_chart = productos_gasto_list[:10]
        chart_labels  = [p['producto'].nombre for p in top_chart]
        chart_values  = [p['total_valorizado'] for p in top_chart]
        chart_quantities = [p['cantidad_total'] for p in top_chart]
        chart_prices     = [p['precio_unitario'] for p in top_chart]

        context.update({
            'es_detalle': True,
            'modulo_nombre': modulo_nombre,
            'productos_gasto': page_obj,
            'total_modulo': total_modulo,
            'total_cantidades': total_cantidades,
            'page_obj': page_obj,
            'paginator': paginator,
            'chart_labels': chart_labels,
            'chart_values': chart_values,
            'chart_quantities': chart_quantities,
            'chart_prices': chart_prices,
        })
    else:
        gasto_por_modulo = {}
        for m in modulos:
            gasto_por_modulo[m.pk] = {
                'modulo': m,
                'modulo_id': m.pk,
                'nombre': m.nombre,
                'cantidad_total': 0,
                'total_valorizado': 0.0
            }
        gasto_por_modulo['sin_modulo'] = {
            'modulo': None,
            'modulo_id': 'sin_modulo',
            'nombre': 'Sin Módulo / Torre',
            'cantidad_total': 0,
            'total_valorizado': 0.0
        }
        
        for d in salidas_det:
            m_id = d.salida.modulo_torre_id if d.salida.modulo_torre_id else 'sin_modulo'
            prod_id = d.producto_id
            precio_unitario = precios_dict.get(prod_id, 0.0)
            subtotal = float(d.cantidad) * precio_unitario
            if m_id in gasto_por_modulo:
                gasto_por_modulo[m_id]['cantidad_total'] += d.cantidad
                gasto_por_modulo[m_id]['total_valorizado'] += subtotal
                
        gasto_list = list(gasto_por_modulo.values())
        gasto_list.sort(key=lambda x: x['total_valorizado'], reverse=True)
        
        total_general = sum(g['total_valorizado'] for g in gasto_list)
        total_cantidades_general = sum(g['cantidad_total'] for g in gasto_list)
        
        paginator = Paginator(gasto_list, 50)
        page = request.GET.get('page', 1)
        page_obj = paginator.get_page(page)
        
        # Datos para gráfico de dona (módulos con gasto > 0)
        gasto_list_con_gasto = [g for g in gasto_list if g['total_valorizado'] > 0]
        chart_labels_gen  = [g['nombre'] for g in gasto_list_con_gasto]
        chart_values_gen  = [g['total_valorizado'] for g in gasto_list_con_gasto]
        chart_qtys_gen    = [g['cantidad_total'] for g in gasto_list_con_gasto]

        context.update({
            'es_detalle': False,
            'gastos': page_obj,
            'total_general': total_general,
            'total_cantidades_general': total_cantidades_general,
            'page_obj': page_obj,
            'paginator': paginator,
            'chart_labels_gen': chart_labels_gen,
            'chart_values_gen': chart_values_gen,
            'chart_qtys_gen': chart_qtys_gen,
        })
        
    return render(request, 'inventory/reporte_gasto_modulo_list.html', context)


@login_required
@permission_required('inventory.view_salida', raise_exception=True)
def reporte_gasto_modulo_torre_pdf(request):
    modulo_id = request.GET.get('modulo')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    salidas_det = DetalleSalida.objects.select_related('salida__modulo_torre', 'producto')
    
    if request.user.proyecto and not request.user.es_admin:
        salidas_det = salidas_det.filter(salida__proyecto=request.user.proyecto)
        
    if modulo_id:
        if modulo_id == 'sin_modulo':
            salidas_det = salidas_det.filter(salida__modulo_torre__isnull=True)
        else:
            salidas_det = salidas_det.filter(salida__modulo_torre_id=modulo_id)
            
    if fecha_inicio:
        salidas_det = salidas_det.filter(salida__fecha__gte=fecha_inicio)
    if fecha_fin:
        salidas_det = salidas_det.filter(salida__fecha__lte=fecha_fin)
        
    precios_dict = {}
    ultimos_ingresos = DetalleIngreso.objects.values('producto').annotate(max_id=Max('idDetalle'))
    ids_ingresos = [i['max_id'] for i in ultimos_ingresos]
    precios_qs = DetalleIngreso.objects.filter(idDetalle__in=ids_ingresos).values('producto_id', 'precio')
    for p in precios_qs:
        precios_dict[p['producto_id']] = float(p['precio'])
        
    empresa = Empresa.objects.first()
    titulo_reporte = 'Reporte de Gasto de Material Valorizado'
    
    filtros = []
    if fecha_inicio:
        filtros.append(f'Desde: {fecha_inicio}')
    if fecha_fin:
        filtros.append(f'Hasta: {fecha_fin}')
        
    if modulo_id:
        if modulo_id == 'sin_modulo':
            modulo_nombre = 'Sin Módulo / Torre'
        else:
            modulo_obj = get_object_or_404(ModuloTorre, pk=modulo_id)
            modulo_nombre = modulo_obj.nombre
        filtros.append(f'Módulo: {modulo_nombre}')
        
        if filtros:
            titulo_reporte += f" ({', '.join(filtros)})"
            
        productos_gasto = {}
        for d in salidas_det:
            prod_id = d.producto_id
            precio_unitario = precios_dict.get(prod_id, 0.0)
            subtotal = float(d.cantidad) * precio_unitario
            
            if prod_id not in productos_gasto:
                productos_gasto[prod_id] = {
                    'producto': d.producto,
                    'cantidad_total': 0,
                    'precio_unitario': precio_unitario,
                    'total_valorizado': 0.0
                }
            productos_gasto[prod_id]['cantidad_total'] += d.cantidad
            productos_gasto[prod_id]['total_valorizado'] += subtotal
            
        productos_gasto_list = sorted(productos_gasto.values(), key=lambda x: x['total_valorizado'], reverse=True)
        total_modulo = sum(p['total_valorizado'] for p in productos_gasto_list)
        total_cantidades = sum(p['cantidad_total'] for p in productos_gasto_list)
        
        context = {
            'es_detalle': True,
            'modulo_nombre': modulo_nombre,
            'productos_gasto': productos_gasto_list,
            'total_modulo': total_modulo,
            'total_cantidades': total_cantidades,
            'titulo_reporte': titulo_reporte,
            'empresa': empresa,
        }
    else:
        if filtros:
            titulo_reporte += f" ({', '.join(filtros)})"
            
        modulos = ModuloTorre.objects.all().order_by('nombre')
        gasto_por_modulo = {}
        for m in modulos:
            gasto_por_modulo[m.pk] = {
                'nombre': m.nombre,
                'cantidad_total': 0,
                'total_valorizado': 0.0
            }
        gasto_por_modulo['sin_modulo'] = {
            'nombre': 'Sin Módulo / Torre',
            'cantidad_total': 0,
            'total_valorizado': 0.0
        }
        
        for d in salidas_det:
            m_id = d.salida.modulo_torre_id if d.salida.modulo_torre_id else 'sin_modulo'
            prod_id = d.producto_id
            precio_unitario = precios_dict.get(prod_id, 0.0)
            subtotal = float(d.cantidad) * precio_unitario
            if m_id in gasto_por_modulo:
                gasto_por_modulo[m_id]['cantidad_total'] += d.cantidad
                gasto_por_modulo[m_id]['total_valorizado'] += subtotal
                
        gasto_list = sorted(gasto_por_modulo.values(), key=lambda x: x['total_valorizado'], reverse=True)
        total_general = sum(g['total_valorizado'] for g in gasto_list)
        total_cantidades_general = sum(g['cantidad_total'] for g in gasto_list)
        
        context = {
            'es_detalle': False,
            'gastos': gasto_list,
            'total_general': total_general,
            'total_cantidades_general': total_cantidades_general,
            'titulo_reporte': titulo_reporte,
            'empresa': empresa,
        }
        
    return render(request, 'inventory/reportes/gasto_modulo.html', context)


@login_required
@permission_required('inventory.view_salida', raise_exception=True)
def exportar_gasto_modulo_torre_excel(request):
    modulo_id = request.GET.get('modulo')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    salidas_det = DetalleSalida.objects.select_related('salida__modulo_torre', 'producto')
    
    if request.user.proyecto and not request.user.es_admin:
        salidas_det = salidas_det.filter(salida__proyecto=request.user.proyecto)
        
    if modulo_id:
        if modulo_id == 'sin_modulo':
            salidas_det = salidas_det.filter(salida__modulo_torre__isnull=True)
        else:
            salidas_det = salidas_det.filter(salida__modulo_torre_id=modulo_id)
            
    if fecha_inicio:
        salidas_det = salidas_det.filter(salida__fecha__gte=fecha_inicio)
    if fecha_fin:
        salidas_det = salidas_det.filter(salida__fecha__lte=fecha_fin)
        
    precios_dict = {}
    ultimos_ingresos = DetalleIngreso.objects.values('producto').annotate(max_id=Max('idDetalle'))
    ids_ingresos = [i['max_id'] for i in ultimos_ingresos]
    precios_qs = DetalleIngreso.objects.filter(idDetalle__in=ids_ingresos).values('producto_id', 'precio')
    for p in precios_qs:
        precios_dict[p['producto_id']] = float(p['precio'])
        
    wb = Workbook()
    ws = wb.active
    
    if modulo_id:
        if modulo_id == 'sin_modulo':
            modulo_nombre = 'Sin Módulo / Torre'
        else:
            modulo_obj = get_object_or_404(ModuloTorre, pk=modulo_id)
            modulo_nombre = modulo_obj.nombre
            
        ws.title = f"Gasto {modulo_nombre[:20]}"
        ws.append([f'Reporte Detallado de Gasto - Módulo/Torre: {modulo_nombre}'])
        if fecha_inicio or fecha_fin:
            ws.append([f'Período: {fecha_inicio or "Inicio"} - {fecha_fin or "Fin"}'])
        ws.append([])
        
        ws.append(['Producto', 'Precio Unitario', 'Cantidad Total', 'Total Valorizado'])
        
        productos_gasto = {}
        for d in salidas_det:
            prod_id = d.producto_id
            precio_unitario = precios_dict.get(prod_id, 0.0)
            subtotal = float(d.cantidad) * precio_unitario
            if prod_id not in productos_gasto:
                productos_gasto[prod_id] = {
                    'nombre': d.producto.nombre,
                    'cantidad_total': 0,
                    'precio_unitario': precio_unitario,
                    'total_valorizado': 0.0
                }
            productos_gasto[prod_id]['cantidad_total'] += d.cantidad
            productos_gasto[prod_id]['total_valorizado'] += subtotal
            
        productos_gasto_list = sorted(productos_gasto.values(), key=lambda x: x['total_valorizado'], reverse=True)
        for p in productos_gasto_list:
            ws.append([p['nombre'], p['precio_unitario'], p['cantidad_total'], p['total_valorizado']])
            
        ws.append([])
        total_modulo = sum(p['total_valorizado'] for p in productos_gasto_list)
        total_cantidades = sum(p['cantidad_total'] for p in productos_gasto_list)
        ws.append(['TOTAL', '', total_cantidades, total_modulo])
    else:
        ws.title = "Gasto General Módulos"
        ws.append(['Reporte Resumido de Gasto por Módulo/Torre'])
        if fecha_inicio or fecha_fin:
            ws.append([f'Período: {fecha_inicio or "Inicio"} - {fecha_fin or "Fin"}'])
        ws.append([])
        
        ws.append(['Módulo/Torre', 'Cantidad Total Material', 'Total Valorizado'])
        
        modulos = ModuloTorre.objects.all().order_by('nombre')
        gasto_por_modulo = {}
        for m in modulos:
            gasto_por_modulo[m.pk] = {
                'nombre': m.nombre,
                'cantidad_total': 0,
                'total_valorizado': 0.0
            }
        gasto_por_modulo['sin_modulo'] = {
            'nombre': 'Sin Módulo / Torre',
            'cantidad_total': 0,
            'total_valorizado': 0.0
        }
        
        for d in salidas_det:
            m_id = d.salida.modulo_torre_id if d.salida.modulo_torre_id else 'sin_modulo'
            prod_id = d.producto_id
            precio_unitario = precios_dict.get(prod_id, 0.0)
            subtotal = float(d.cantidad) * precio_unitario
            if m_id in gasto_por_modulo:
                gasto_por_modulo[m_id]['cantidad_total'] += d.cantidad
                gasto_por_modulo[m_id]['total_valorizado'] += subtotal
                
        gasto_list = sorted(gasto_por_modulo.values(), key=lambda x: x['total_valorizado'], reverse=True)
        for g in gasto_list:
            ws.append([g['nombre'], g['cantidad_total'], g['total_valorizado']])
            
        ws.append([])
        total_general = sum(g['total_valorizado'] for g in gasto_list)
        total_cantidades_general = sum(g['cantidad_total'] for g in gasto_list)
        ws.append(['TOTAL GENERAL', total_cantidades_general, total_general])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=gasto_material_modulos.xlsx'
    wb.save(response)
    return response



@login_required
def reporte_movimientos_usuario_list_view(request):
    from ..models import HistorialMovimiento, Usuario

    usuario_id = request.GET.get('usuario_id', '')
    tipo       = request.GET.get('tipo', '')
    desde      = request.GET.get('desde', '')
    hasta      = request.GET.get('hasta', '')

    qs = HistorialMovimiento.objects.select_related('usuario', 'bodega').order_by('-fecha')

    if usuario_id:
        qs = qs.filter(usuario_id=usuario_id)
    if tipo:
        qs = qs.filter(tipo_accion=tipo)
    if desde:
        qs = qs.filter(fecha__date__gte=desde)
    if hasta:
        qs = qs.filter(fecha__date__lte=hasta)

    # Resumen de conteos por tipo para el usuario seleccionado
    from django.db.models import Count
    resumen = (
        qs.values('tipo_accion')
          .annotate(total=Count('id'))
          .order_by('tipo_accion')
    ) if usuario_id else []

    paginator  = Paginator(qs, 50)
    page_obj   = paginator.get_page(request.GET.get('page'))
    usuarios   = Usuario.objects.filter(is_active=True).order_by('username')
    usuario_sel = None
    if usuario_id:
        try:
            usuario_sel = Usuario.objects.get(pk=usuario_id)
        except Usuario.DoesNotExist:
            pass

    tipo_label = {k: v for k, v in HistorialMovimiento.TIPO_CHOICES}

    return render(request, 'inventory/reporte_movimientos_usuario.html', {
        'page_obj':         page_obj,
        'usuarios':         usuarios,
        'usuario_sel':      usuario_sel,
        'resumen':          resumen,
        'tipo_label':       tipo_label,
        'tipo_choices':     HistorialMovimiento.TIPO_CHOICES,
        'filtro_usuario_id': usuario_id,
        'filtro_tipo':      tipo,
        'filtro_desde':     desde,
        'filtro_hasta':     hasta,
    })


def _build_gastos_qs(request):
    """Aplica los filtros comunes de la vista de gastos al queryset."""
    qs = Gasto.objects.select_related('trabajador', 'proyecto').filter(activo=True)
    concepto   = request.GET.get('concepto')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    proyecto    = request.GET.get('proyecto')

    if concepto:
        qs = qs.filter(concepto=concepto)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)
    if proyecto:
        qs = qs.filter(proyecto_id=proyecto)
    return qs


@login_required
@permission_required('inventory.view_gasto', raise_exception=True)
def reporte_gastos_list_view(request):
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    context = {
        'proyectos': proyectos,
        'conceptos': Gasto.CONCEPTO_CHOICES,
        'fecha_desde': request.GET.get('fecha_desde', ''),
        'fecha_hasta': request.GET.get('fecha_hasta', ''),
        'proyecto_id': request.GET.get('proyecto', ''),
        'concepto': request.GET.get('concepto', ''),
    }
    return render(request, 'inventory/reporte_gastos_list.html', context)


@login_required
@permission_required('inventory.view_gasto', raise_exception=True)
def reporte_gastos_pdf(request):
    gastos = _build_gastos_qs(request)
    total = gastos.aggregate(total=Sum('monto'))['total'] or 0

    filtros = []
    if request.GET.get('proyecto'):
        proj = Proyecto.objects.filter(pk=request.GET['proyecto']).first()
        if proj:
            filtros.append(f'Proyecto: {proj.nombre}')
    if request.GET.get('concepto'):
        concepto_label = dict(Gasto.CONCEPTO_CHOICES).get(request.GET['concepto'], request.GET['concepto'])
        filtros.append(f'Concepto: {concepto_label}')
    if request.GET.get('fecha_desde'):
        filtros.append(f'Desde: {request.GET["fecha_desde"]}')
    if request.GET.get('fecha_hasta'):
        filtros.append(f'Hasta: {request.GET["fecha_hasta"]}')

    titulo_reporte = 'Reporte de Gastos'
    if filtros:
        titulo_reporte += f" ({', '.join(filtros)})"

    empresa = Empresa.objects.first()
    return render(request, 'inventory/reportes/gastos.html', {
        'gastos': gastos,
        'total': total,
        'titulo_reporte': titulo_reporte,
        'empresa': empresa,
        'filtros': filtros,
    })


@login_required
@permission_required('inventory.view_gasto', raise_exception=True)
def exportar_gastos_excel(request):
    from openpyxl.styles import Font, PatternFill, Alignment
    gastos = _build_gastos_qs(request)
    total = gastos.aggregate(total=Sum('monto'))['total'] or 0

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gastos'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3A8A')
    total_fill  = PatternFill('solid', fgColor='DBEAFE')
    total_font  = Font(bold=True, color='1E3A8A')
    center      = Alignment(horizontal='center')

    headers = ['Fecha', 'Tipo Doc.', 'N° Documento', 'Concepto', 'Observaciones', 'Monto ($)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    concepto_display  = dict(Gasto.CONCEPTO_CHOICES)
    tipo_doc_display  = dict(Gasto.TIPO_DOC_CHOICES)
    for g in gastos:
        ws.append([
            g.fecha.strftime('%d/%m/%Y') if g.fecha else '',
            tipo_doc_display.get(g.tipo_documento, '') if g.tipo_documento else '',
            g.num_documento or '',
            concepto_display.get(g.concepto, g.concepto),
            g.observaciones or '',
            float(g.monto),
        ])

    total_row = ws.max_row + 1
    ws.cell(total_row, 5, 'TOTAL').font      = total_font
    ws.cell(total_row, 5).fill               = total_fill
    ws.cell(total_row, 5).alignment          = center
    ws.cell(total_row, 6, float(total)).font = total_font
    ws.cell(total_row, 6).fill               = total_fill

    col_widths = [12, 14, 20, 16, 38, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=gastos.xlsx'
    wb.save(response)
    return response
